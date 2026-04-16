from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
import io

from app.core.database import get_db
from app.core.security import get_current_user, require_role
from app.models.models import (
    Timetable, TimetableEntry, User, UserRole,
    Subject, Teacher, Section, Room, TimeSlot, Department,
)
from app.schemas.schemas import (
    TimetableCreate, TimetableResponse, TimetableEntryResponse,
    TimetableEntryUpdate, GenerateRequest,
)
from app.services.scheduler import TimetableScheduler, detect_conflicts, suggest_improvements

router = APIRouter(prefix="/api/timetables", tags=["timetables"])


def _enrich_entry(entry: TimetableEntry, db: Session) -> dict:
    """Add nested display data to a timetable entry."""
    subject = db.query(Subject).filter(Subject.id == entry.subject_id).first()
    teacher = db.query(Teacher).filter(Teacher.id == entry.teacher_id).first()
    section = db.query(Section).filter(Section.id == entry.section_id).first()
    room = db.query(Room).filter(Room.id == entry.room_id).first() if entry.room_id else None
    timeslot = db.query(TimeSlot).filter(TimeSlot.id == entry.timeslot_id).first()

    return {
        "id": entry.id,
        "timetable_id": entry.timetable_id,
        "timeslot_id": entry.timeslot_id,
        "subject_id": entry.subject_id,
        "teacher_id": entry.teacher_id,
        "section_id": entry.section_id,
        "room_id": entry.room_id,
        "subject_name": subject.name if subject else None,
        "subject_code": subject.code if subject else None,
        "subject_color": subject.color if subject else None,
        "teacher_name": teacher.name if teacher else None,
        "teacher_color": teacher.color if teacher else None,
        "section_name": section.name if section else None,
        "room_name": room.name if room else None,
        "day": timeslot.day.value if timeslot and hasattr(timeslot.day, 'value') else (timeslot.day if timeslot else None),
        "period": timeslot.period if timeslot else None,
        "start_time": timeslot.start_time if timeslot else None,
        "end_time": timeslot.end_time if timeslot else None,
        "is_lab": subject.is_lab if subject else False,
    }


@router.get("", response_model=List[TimetableResponse])
def list_timetables(
    department_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(Timetable)
    if department_id:
        query = query.filter(Timetable.department_id == department_id)
    timetables = query.order_by(Timetable.created_at.desc()).all()

    result = []
    for tt in timetables:
        tt_dict = {
            "id": tt.id,
            "name": tt.name,
            "department_id": tt.department_id,
            "academic_year": tt.academic_year,
            "semester": tt.semester,
            "is_published": tt.is_published,
            "score": tt.score,
            "generation_log": tt.generation_log,
            "created_at": tt.created_at,
            "updated_at": tt.updated_at,
            "entries": None,
        }
        result.append(tt_dict)
    return result


@router.get("/{timetable_id}")
def get_timetable(timetable_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    tt = db.query(Timetable).filter(Timetable.id == timetable_id).first()
    if not tt:
        raise HTTPException(status_code=404, detail="Timetable not found")

    entries = db.query(TimetableEntry).filter(TimetableEntry.timetable_id == timetable_id).all()
    enriched_entries = [_enrich_entry(e, db) for e in entries]

    return {
        "id": tt.id,
        "name": tt.name,
        "department_id": tt.department_id,
        "academic_year": tt.academic_year,
        "semester": tt.semester,
        "is_published": tt.is_published,
        "score": tt.score,
        "generation_log": tt.generation_log,
        "created_at": tt.created_at,
        "updated_at": tt.updated_at,
        "entries": enriched_entries,
    }


@router.post("/generate")
def generate_timetable(
    data: GenerateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.SUPER_ADMIN, UserRole.ADMIN)),
):
    dept = db.query(Department).filter(Department.id == data.department_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")

    scheduler = TimetableScheduler(db, data.department_id, dept.college_id)
    result = scheduler.generate()

    if not result.success:
        return {
            "success": False,
            "log": result.log,
            "message": "Failed to generate timetable. Check logs for details.",
        }

    # Save timetable
    timetable = Timetable(
        name=data.name or f"Timetable - {dept.name}",
        department_id=data.department_id,
        academic_year=data.academic_year,
        semester=data.semester,
        score=result.score,
        generation_log={"log": result.log},
    )
    db.add(timetable)
    db.commit()
    db.refresh(timetable)

    # Save entries
    for entry_data in result.entries:
        entry = TimetableEntry(
            timetable_id=timetable.id,
            timeslot_id=entry_data["timeslot_id"],
            subject_id=entry_data["subject_id"],
            teacher_id=entry_data["teacher_id"],
            section_id=entry_data["section_id"],
            room_id=entry_data.get("room_id"),
        )
        db.add(entry)
    db.commit()

    return {
        "success": True,
        "timetable_id": timetable.id,
        "score": result.score,
        "entries_count": len(result.entries),
        "log": result.log,
    }


@router.put("/{timetable_id}/entries/{entry_id}")
def update_entry(
    timetable_id: int,
    entry_id: int,
    data: TimetableEntryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.SUPER_ADMIN, UserRole.ADMIN)),
):
    entry = db.query(TimetableEntry).filter(
        TimetableEntry.id == entry_id,
        TimetableEntry.timetable_id == timetable_id,
    ).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    entry.timeslot_id = data.timeslot_id
    entry.subject_id = data.subject_id
    entry.teacher_id = data.teacher_id
    entry.section_id = data.section_id
    entry.room_id = data.room_id
    db.commit()
    db.refresh(entry)
    return _enrich_entry(entry, db)


@router.delete("/{timetable_id}/entries/{entry_id}")
def delete_entry(
    timetable_id: int,
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.SUPER_ADMIN, UserRole.ADMIN)),
):
    entry = db.query(TimetableEntry).filter(
        TimetableEntry.id == entry_id,
        TimetableEntry.timetable_id == timetable_id,
    ).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    db.delete(entry)
    db.commit()
    return {"detail": "Entry deleted"}


@router.post("/{timetable_id}/publish")
def publish_timetable(
    timetable_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.SUPER_ADMIN, UserRole.ADMIN)),
):
    tt = db.query(Timetable).filter(Timetable.id == timetable_id).first()
    if not tt:
        raise HTTPException(status_code=404, detail="Timetable not found")
    tt.is_published = True
    db.commit()
    return {"detail": "Timetable published"}


@router.get("/{timetable_id}/conflicts")
def get_conflicts(
    timetable_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    conflicts = detect_conflicts(db, timetable_id)
    return {"conflicts": conflicts, "count": len(conflicts)}


@router.get("/{timetable_id}/suggestions")
def get_suggestions(
    timetable_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    suggestions = suggest_improvements(db, timetable_id)
    return {"suggestions": suggestions, "count": len(suggestions)}


@router.delete("/{timetable_id}")
def delete_timetable(
    timetable_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.SUPER_ADMIN, UserRole.ADMIN)),
):
    tt = db.query(Timetable).filter(Timetable.id == timetable_id).first()
    if not tt:
        raise HTTPException(status_code=404, detail="Timetable not found")
    db.delete(tt)
    db.commit()
    return {"detail": "Timetable deleted"}


@router.get("/{timetable_id}/export/excel")
def export_excel(
    timetable_id: int,
    section_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tt = db.query(Timetable).filter(Timetable.id == timetable_id).first()
    if not tt:
        raise HTTPException(status_code=404, detail="Timetable not found")

    entries = db.query(TimetableEntry).filter(TimetableEntry.timetable_id == timetable_id)
    if section_id:
        entries = entries.filter(TimetableEntry.section_id == section_id)
    entries = entries.all()

    wb = Workbook()
    ws = wb.active
    ws.title = tt.name[:31]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    days = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday"]
    periods = sorted(set(
        db.query(TimeSlot.period).filter(TimeSlot.college_id == tt.department.college_id).all()
    ))
    period_nums = [p[0] for p in periods]

    # Write headers
    ws.cell(row=1, column=1, value="Day/Period").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = thin_border

    for i, period in enumerate(period_nums):
        slot = db.query(TimeSlot).filter(
            TimeSlot.period == period,
            TimeSlot.college_id == tt.department.college_id,
        ).first()
        header_text = f"P{period}\n{slot.start_time}-{slot.end_time}" if slot else f"P{period}"
        cell = ws.cell(row=1, column=i + 2, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Write timetable data
    entry_map = {}
    for e in entries:
        enriched = _enrich_entry(e, db)
        key = (enriched["day"], enriched["period"])
        entry_map[key] = enriched

    for row_idx, day in enumerate(days, start=2):
        cell = ws.cell(row=row_idx, column=1, value=day.capitalize())
        cell.font = Font(bold=True)
        cell.border = thin_border

        for col_idx, period in enumerate(period_nums, start=2):
            entry = entry_map.get((day, period))
            if entry:
                text = f"{entry['subject_code']}\n{entry['teacher_name']}"
                if entry.get('room_name'):
                    text += f"\n{entry['room_name']}"
            else:
                text = "-"
            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 15
    for i in range(len(period_nums)):
        col_letter = chr(66 + i)
        ws.column_dimensions[col_letter].width = 18

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={tt.name}.xlsx"},
    )
